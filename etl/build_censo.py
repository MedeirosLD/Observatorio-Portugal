# -*- coding: utf-8 -*-
"""
Constrói dados/censo/*.json a partir de fontes oficiais do INE:

  - Censos 2021: BGRI ficheiro síntese (subsecção)  -> população, sexo,
    faixas etárias, agregados, habitação (freguesia + concelho)
  - Censos 2021: indicadores API INE                -> taxa de desemprego,
    taxa de atividade, % ensino superior
  - Censos 2011: BGRI geopackage                    -> mesmos campos + emprego
    e escolaridade calculados do próprio ficheiro
  - Censos 2001: BGRI (quando disponível/na mesma estrutura)
  - Rendimento (IRS): valor mediano do rendimento bruto declarado deduzido
    do IRS liquidado por agregado fiscal, por freguesia (série 2018-2023)

Uso:
  python etl/build_censo.py --raw <pasta-com-downloads> [--out dados/censo]

A pasta raw deve conter (baixados de mapas.ine.pt):
  fs2021/FS2021_subseccao.xlsx   (FS2021SubSeccaoTot.zip)
  portugal2011.gpkg              (filesGPG/2011/portugal2011.zip)
  portugal2001.gpkg              (filesGPG/2001/portugal2001.zip, opcional)
Os indicadores da API INE são baixados diretamente (requer rede).
"""

import argparse
import json
import sqlite3
import sys
import urllib.request
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent

INE_API = "https://www.ine.pt/ine/json_indicador/pindica.jsp?op=2&varcd={var}&Dim1=S7A{ano}&lang=PT"

# Indicadores Censos 2021 (por freguesia; Dim3=Sexo quando existir -> usar total 'T')
IND_2021 = {
    "tx_desemprego": "0012328",
    "tx_atividade": "0012317",
    "pct_ens_superior": "0012316",
    # escolaridade (proporções oficiais, cada uma com a sua base)
    "esc_sem": "0012323",      # 15+ sem nenhum nível completo
    "esc_3ciclo": "0012315",   # pelo menos 3º ciclo do ensino básico
    "esc_sec": "0012327",      # pelo menos ensino secundário
}
# População residente por grupo etário decenal (0012355): Dim4 01..11
IND_IDADE_DEC = "0012355"
IDADE_DEC_GRUPOS = {
    "01": "d0_9", "02": "d10_19", "03": "d20_29", "04": "d30_39",
    "05": "d40_49", "06": "d50_59", "07": "d60_69", "08": "d70_79",
    "09": "d80", "10": "d80", "11": "d80",  # 80-89/90-99/100+ agregados em 80+
}
IND_RENDA = "0012741"  # mediana rendimento bruto - IRS liquidado, por agregado fiscal
ANOS_RENDA = [2018, 2019, 2020, 2021, 2022, 2023]


def fetch_ine(var, ano, extra=""):
    import time
    url = INE_API.format(var=var, ano=ano) + extra
    for tent in range(5):
        try:
            with urllib.request.urlopen(url, timeout=120) as r:
                payload = json.load(r)
            dados = payload[0].get("Dados") or {}
            key = f"S7A{ano}"
            return dados.get(key) or next(iter(dados.values()), [])
        except Exception:
            if tent == 4:
                raise
            time.sleep(20 * (tent + 1))  # rate limit do INE


def geocod_to_dicofre(geocod):
    """Extrai o dicofre (6 díg.) / dico (4 díg.) do geocod INE (NUTS-2024).

    Formato: NUTS3 (3 chars, pode ter letras) + código administrativo.
    Ex.: '1C1021111' (9) -> freguesia 021111; '1C10211' (7) -> concelho 0211;
         '300310905' (9) -> freguesia 310905 (RA Madeira); '1701106' (7) ->
         concelho 1106 (Lisboa, AML='170').
    Devolve (nivel, codigo) ou (None, None) quando não é freguesia/concelho.
    """
    g = str(geocod)
    if len(g) == 6 and g.isdigit():
        return "freg", g
    if len(g) == 4 and g.isdigit():
        return "conc", g
    if len(g) == 9 and g[-6:].isdigit():
        return "freg", g[-6:]
    if len(g) == 7 and g[-4:].isdigit():
        return "conc", g[-4:]
    return None, None


def parse_ine_rows(rows):
    """rows da API -> ({dicofre: valor}, {dico: valor})"""
    freg, conc = {}, {}
    for row in rows:
        val = row.get("valor")
        if val in (None, ""):
            continue
        try:
            v = float(val)
        except ValueError:
            continue
        nivel, cod = geocod_to_dicofre(row.get("geocod", ""))
        if nivel == "freg":
            freg[cod] = v
        elif nivel == "conc":
            conc[cod] = v
    return freg, conc


# ---------------------------------------------------------------- 2021 (BGRI)

COLS_2021 = {
    "pop": "N_INDIVIDUOS",
    "pop_h": "N_INDIVIDUOS_H",
    "pop_m": "N_INDIVIDUOS_M",
    "i0_14": "N_INDIVIDUOS_0_14",
    "i15_24": "N_INDIVIDUOS_15_24",
    "i25_64": "N_INDIVIDUOS_25_64",
    "i65": "N_INDIVIDUOS_65_OU_MAIS",
    "agregados": "N_AGREGADOS_DOMESTICOS_PRIVADOS",
    "aloj": "N_ALOJAMENTOS_FAMILIARES",
    "aloj_rhab": "N_ALOJAMENTOS_FAM_CLASS_RHABITUAL",
    "aloj_prop": "N_RHABITUAL_PROP_OCUP",
    "aloj_arren": "N_RHABITUAL_ARRENDADOS",
    "edif": "N_EDIFICIOS_CLASSICOS",
}


def build_2021(xlsx_path):
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=2, values_only=True)
    hdr = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(hdr)}
    col_mun, col_fr = idx["MUNICIPIO"], idx["FREGUESIA"]
    col_sec, col_ss = idx["SECCAO"], idx["SUBSECCAO"]
    picks = {k: idx[c] for k, c in COLS_2021.items()}

    freg, conc = {}, {}
    nacional = None
    for r in rows:
        mun, fr = r[col_mun], r[col_fr]
        sec, ss = r[col_sec], r[col_ss]
        if sec not in (None, "", "Total") or ss not in (None, "", "Total"):
            continue  # só totais (freguesia/município/superiores)
        rec = {}
        for k, i in picks.items():
            v = r[i]
            rec[k] = int(v) if isinstance(v, (int, float)) and v == v else 0
        if fr not in (None, "", "Total"):
            freg[str(fr).zfill(6)] = rec
        elif mun not in (None, "", "Total"):
            conc[str(mun).zfill(4)] = rec
        elif nacional is None and rec.get("pop"):
            nacional = rec  # primeira linha total (Portugal)
    return freg, conc, nacional


# ---------------------------------------------------------- 2011/2001 (GPKG)

def gpkg_tables(con, prefix):
    q = ("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE ? "
         "AND name NOT LIKE 'gpkg%'")
    return [r[0] for r in con.execute(q, (prefix + "%",))]


def sum_expr(cols, alias):
    return ", ".join(f"SUM(COALESCE(\"{c}\", 0)) AS {a}" for a, c in cols.items()) or ""


COLS_2011 = {
    "pop": "N_INDIVIDUOS_RESIDENT",
    "pop_h": "N_INDIVIDUOS_RESIDENT_H",
    "pop_m": "N_INDIVIDUOS_RESIDENT_M",
    "i0_4": "N_INDIVIDUOS_RESIDENT_0A4",
    "i5_9": "N_INDIVIDUOS_RESIDENT_5A9",
    "i10_13": "N_INDIVIDUOS_RESIDENT_10A13",
    "i14_19": "N_INDIVIDUOS_RESIDENT_14A19",
    "i15_19": "N_INDIVIDUOS_RESIDENT_15A19",
    "i20_24": "N_INDIVIDUOS_RESIDENT_20A24",
    "i25_64": "N_INDIVIDUOS_RESIDENT_25A64",
    "i65": "N_INDIVIDUOS_RESIDENT_65",
    "agregados": "N_FAMILIAS_CLASSICAS",
    "aloj": "N_ALOJAMENTOS_FAMILIARES",
    "aloj_rhab": "N_ALOJAMENTOS_RES_HABITUAL",
    "edif": "N_EDIFICIOS_CLASSICOS",
    "empregados": "N_IND_RESID_EMPREGADOS",
    "desemp_1emp": "N_IND_RESID_DESEMP_PROC_1EMPRG",
    "desemp_novo": "N_IND_RESID_DESEMP_PROC_EMPRG",
    "ens_sup": "N_IND_RESIDENT_ENSINCOMP_SUP",
    "ens_sec": "N_IND_RESIDENT_ENSINCOMP_SEC",
    "ens_posec": "N_IND_RESIDENT_ENSINCOMP_POSEC",
    "ens_3bas": "N_IND_RESIDENT_ENSINCOMP_3BAS",
}


def build_gpkg(gpkg_path, table_prefix, key_cols, cols):
    """Agrega SUM(cols) por freguesia. key_cols: colunas concatenadas que
    formam o dicofre (ex.: ('DTMN11','FR11') ou ('DTCCFR01',))."""
    con = sqlite3.connect(gpkg_path)
    freg = {}
    seen_tabs = set()
    for tab in gpkg_tables(con, table_prefix):
        tcols = {r[1] for r in con.execute(f'PRAGMA table_info("{tab}")')}
        if any(k not in tcols for k in key_cols):
            continue
        # evita tabelas duplicadas (ex.: BGRI_01CONT vs BGRI01_CONT com o mesmo conteúdo)
        sig = (con.execute(f'SELECT COUNT(*) FROM "{tab}"').fetchone()[0])
        if sig in seen_tabs:
            continue
        seen_tabs.add(sig)
        use = {a: c for a, c in cols.items() if c in tcols}
        if not use:
            continue
        key_expr = " || ".join(f'"{k}"' for k in key_cols)
        sel = sum_expr(use, "x")
        q = (f'SELECT {key_expr} AS dicofre, {sel} '
             f'FROM "{tab}" GROUP BY {key_expr}')
        names = ["dicofre"] + list(use.keys())
        for row in con.execute(q):
            rec = dict(zip(names, row))
            code = str(rec.pop("dicofre")).zfill(6)
            acc = freg.setdefault(code, {})
            for k, v in rec.items():
                acc[k] = acc.get(k, 0) + int(v or 0)
    con.close()
    return freg


def aggregate_raw(freg_raw):
    """Soma os componentes brutos por concelho (4 primeiros dígitos)."""
    conc = {}
    for code, rec in freg_raw.items():
        acc = conc.setdefault(code[:4], {})
        for k, v in rec.items():
            acc[k] = acc.get(k, 0) + v
    return conc


def finish_2011(raw):
    """Converte contagens 2011 nos mesmos campos do 2021 + emprego/escolaridade."""
    out = {}
    for code, r in raw.items():
        i14_19 = r.get("i14_19", 0)
        i15_19 = r.get("i15_19", 0)
        i14 = max(i14_19 - i15_19, 0)  # os com 14 anos
        rec = {
            "pop": r.get("pop", 0),
            "pop_h": r.get("pop_h", 0),
            "pop_m": r.get("pop_m", 0),
            "i0_14": r.get("i0_4", 0) + r.get("i5_9", 0) + r.get("i10_13", 0) + i14,
            "i15_24": i15_19 + r.get("i20_24", 0),
            "i25_64": r.get("i25_64", 0),
            "i65": r.get("i65", 0),
            "agregados": r.get("agregados", 0),
            "aloj": r.get("aloj", 0),
            "aloj_rhab": r.get("aloj_rhab", 0),
            "edif": r.get("edif", 0),
        }
        emp = r.get("empregados", 0)
        des = r.get("desemp_1emp", 0) + r.get("desemp_novo", 0)
        ativos = emp + des
        if ativos:
            rec["tx_desemprego"] = round(100.0 * des / ativos, 1)
            rec["tx_atividade"] = round(100.0 * ativos / rec["pop"], 1) if rec["pop"] else None
        if rec["pop"]:
            sup = r.get("ens_sup", 0)
            sec_mais = sup + r.get("ens_sec", 0) + r.get("ens_posec", 0)
            c3_mais = sec_mais + r.get("ens_3bas", 0)
            rec["pct_ens_superior"] = round(100.0 * sup / rec["pop"], 1)
            rec["esc_sec"] = round(100.0 * sec_mais / rec["pop"], 1)
            rec["esc_3ciclo"] = round(100.0 * c3_mais / rec["pop"], 1)
        out[code] = {k: v for k, v in rec.items() if v is not None}
    return out


# 2001: ficheiro muito mais pobre — só contagens básicas por subsecção
COLS_2001 = {
    "pop": "RESIDENTES_T",
    "pop_h": "RESIDENTES_H",
    "agregados": "FAMILIAS",
    "aloj": "ALOJAMENTOS",
    "edif": "EDIFICIOS",
}


def finish_2001(raw):
    out = {}
    for code, r in raw.items():
        rec = dict(r)
        if "pop" in rec and "pop_h" in rec:
            rec["pop_m"] = rec["pop"] - rec["pop_h"]
        out[code] = rec
    return out


def write_json(path, obj):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, separators=(",", ":"))
    print(f"  {path.relative_to(ROOT)}  ({path.stat().st_size/1024:.0f} KB)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw", required=True, help="pasta com downloads brutos do INE")
    ap.add_argument("--out", default=str(ROOT / "dados" / "censo"))
    ap.add_argument("--skip-api", action="store_true", help="não consultar a API INE")
    args = ap.parse_args()
    raw = Path(args.raw)
    out = Path(args.out)

    index = {"vintages": [], "renda": None, "fontes": {
        "censos": "INE, Recenseamento da População e Habitação (BGRI)",
        "renda": "INE, Estatísticas do Rendimento ao Nível Local (AT)",
    }}

    # ---- Censos 2021
    xlsx = raw / "fs2021" / "FS2021_subseccao.xlsx"
    if xlsx.exists():
        print("Censos 2021 (BGRI síntese)...")
        freg21, conc21, nac21 = build_2021(xlsx)
        if not args.skip_api:
            print("Indicadores 2021 (API INE)...")
            for campo, var in IND_2021.items():
                rows = fetch_ine(var, 2021)
                fvals, cvals = parse_ine_rows_prefer_total(rows)
                for code, v in fvals.items():
                    if code in freg21:
                        freg21[code][campo] = v
                for code, v in cvals.items():
                    if code in conc21:
                        conc21[code][campo] = v
                print(f"  {campo}: {len(fvals)} freguesias")
            print("Grupos etários decenais 2021 (API INE)...")
            for dim4, campo in IDADE_DEC_GRUPOS.items():
                rows = fetch_ine(IND_IDADE_DEC, 2021, extra=f"&Dim3=T&Dim4={dim4}")
                fvals, cvals = parse_ine_rows_prefer_total(rows)
                for code, v in fvals.items():
                    if code in freg21:
                        idade = freg21[code].setdefault("idade_dec", {})
                        idade[campo] = idade.get(campo, 0) + int(v)
                for code, v in cvals.items():
                    if code in conc21:
                        idade = conc21[code].setdefault("idade_dec", {})
                        idade[campo] = idade.get(campo, 0) + int(v)
                print(f"  grupo {dim4}: {len(fvals)} freguesias")
        write_json(out / "censo_2021.json",
                   {"ano": 2021, "FREG": freg21, "CONC": conc21, "PT": nac21})
        index["vintages"].append(2021)

    # ---- Censos 2011
    gpkg11 = raw / "portugal2011.gpkg"
    if gpkg11.exists():
        print("Censos 2011 (BGRI gpkg)...")
        raw11 = build_gpkg(gpkg11, "BGRI11", ("DTMN11", "FR11"), COLS_2011)
        freg11 = finish_2011(raw11)
        conc11 = finish_2011(aggregate_raw(raw11))
        write_json(out / "censo_2011.json",
                   {"ano": 2011, "FREG": freg11, "CONC": conc11})
        index["vintages"].append(2011)

    # ---- Censos 2001 (ficheiro mais pobre: só contagens básicas)
    gpkg01 = next(iter(raw.glob("*2001*.gpkg")), None)
    if gpkg01:
        print("Censos 2001 (BGRI gpkg)...")
        try:
            raw01 = build_gpkg(gpkg01, "BGRI", ("DTCCFR01",), COLS_2001)
            raw01 = {c: r for c, r in raw01.items() if c.isdigit()}
            if raw01:
                freg01 = finish_2001(raw01)
                conc01 = finish_2001(aggregate_raw(raw01))
                write_json(out / "censo_2001.json",
                           {"ano": 2001, "FREG": freg01, "CONC": conc01})
                index["vintages"].append(2001)
        except Exception as e:
            print(f"  2001 ignorado: {e}")

    # ---- Renda (IRS)
    if not args.skip_api:
        print("Rendimento IRS (API INE)...")
        renda = {"FREG": {}, "CONC": {}, "anos": []}
        for ano in ANOS_RENDA:
            try:
                rows = fetch_ine(IND_RENDA, ano)
            except Exception as e:
                print(f"  renda {ano} falhou: {e}")
                continue
            fvals, cvals = parse_ine_rows(rows)
            if not fvals:
                continue
            renda["anos"].append(ano)
            for code, v in fvals.items():
                renda["FREG"].setdefault(code, {})[str(ano)] = int(v)
            for code, v in cvals.items():
                renda["CONC"].setdefault(code, {})[str(ano)] = int(v)
            print(f"  {ano}: {len(fvals)} freguesias, {len(cvals)} concelhos")
        if renda["anos"]:
            write_json(out / "renda.json", renda)
            index["renda"] = {"anos": renda["anos"], "indicador": IND_RENDA}

    write_json(out / "censo_index.json", index)
    print("OK")


def parse_ine_rows_prefer_total(rows):
    """Como parse_ine_rows, mas em indicadores com dimensão extra (ex.: Sexo)
    fica só com o total (dim_3 == 'T', quando presente)."""
    freg, conc = {}, {}
    for row in rows:
        if row.get("dim_3") not in (None, "T"):
            continue
        val = row.get("valor")
        if val in (None, ""):
            continue
        try:
            v = float(val)
        except ValueError:
            continue
        nivel, cod = geocod_to_dicofre(row.get("geocod", ""))
        if nivel == "freg" and cod not in freg:
            freg[cod] = v
        elif nivel == "conc" and cod not in conc:
            conc[cod] = v
    return freg, conc


if __name__ == "__main__":
    sys.exit(main())
