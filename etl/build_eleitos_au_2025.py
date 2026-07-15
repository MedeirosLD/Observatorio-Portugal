# -*- coding: utf-8 -*-
"""Eleitos das Autárquicas 2025 a partir do mapa oficial retificado (xlsx).

Fonte: mapas_diario_da_republica/autarquicas/2025/mapa_3_eleitos_retificado.xlsx
Estrutura (Folha1): blocos
    CÓD | CONC | FREG | ÓRG      (linha de código; ÓRG in {CM, AM, AF})
    sigla | sigla | ...          (uma coluna por lista)
    nome  | nome  | ...          (nomes na ordem da lista, por coluna)
O primeiro nome da lista mais votada (votos de dados/resultados/au_*.json)
é o presidente da câmara (CM) ou da junta (AF).

Saída: dados/eleitos/au_cm_2025.json, au_am_2025.json, au_af_2025_{dd}.json
"""
import sys
from collections import defaultdict

from openpyxl import load_workbook

from common import norm_dicofre, circulo_from_dicofre, strip_accents_upper
from eleitos_common import (DR_DIR, canonicalize_siglas, compute_presidente,
                            is_person_name, load_results, nfc,
                            norm_name_if_caps, rebuild_index,
                            write_eleitos_json)

XLSX = DR_DIR / "autarquicas" / "2025" / "mapa_3_eleitos_retificado.xlsx"
YEAR = 2025


def raw_blocks(ws):
    """Blocos brutos (cod, conc, freg, orgao, rows) delimitados por linhas de
    código. A linha de código repete-se nas mudanças de página do mapa; blocos
    com o mesmo (código, órgão) são fundidos pela ordem de aparição."""
    order = []
    blocks = {}
    cur = None
    for row in ws.iter_rows(values_only=True):
        cells = [nfc(str(c)).strip() if c is not None and str(c).strip() else None
                 for c in row]
        first = cells[0] if cells else None
        org = cells[3] if len(cells) > 3 else None
        if first and org in ("CM", "AM", "AF") and first[0].isdigit():
            key = (first, org)
            if key not in blocks:
                blocks[key] = (first, cells[1], cells[2], org, [])
                order.append(key)
            cur = blocks[key][4]
            continue
        if cur is None or first == "CÓD" or (first and first.startswith("III")):
            continue
        if any(cells):   # linhas com 1.ª coluna vazia têm nomes nas colunas 2-4
            cur.append(cells)
    for key in order:
        yield blocks[key]


def parse_listas(rows, sigla_slugs):
    """Blocos podem ter vários segmentos "linha de siglas + linhas de nomes"
    (mais de 4 listas não cabem nas 4 colunas). Uma linha é de siglas se todas
    as células não vazias forem siglas conhecidas dos resultados ou não
    parecerem nomes próprios (GCE têm siglas arbitrárias, ex.: "Jonet")."""
    listas = []
    seg = None   # lista de índices em `listas` por coluna
    for cells in rows:
        filled = [c for c in cells if c]
        is_sigla_row = all(
            strip_accents_upper(c) in sigla_slugs or not is_person_name(c)
            for c in filled)
        if is_sigla_row:
            seg = []
            for c in cells:
                if c:
                    listas.append({"sigla": c, "eleitos": []})
                    seg.append(len(listas) - 1)
                else:
                    seg.append(None)
            continue
        if seg is None:
            continue
        for i, c in enumerate(cells):
            if c and i < len(seg) and seg[i] is not None:
                listas[seg[i]]["eleitos"].append(norm_name_if_caps(c))
    # blocos fundidos repetem a linha de siglas: juntar listas da mesma sigla
    merged, by_sigla = [], {}
    for l in listas:
        if not l["eleitos"]:
            continue
        if l["sigla"] in by_sigla:
            by_sigla[l["sigla"]]["eleitos"].extend(l["eleitos"])
        else:
            by_sigla[l["sigla"]] = l
            merged.append(l)
    return merged


def main():
    ws = load_workbook(XLSX, read_only=True)["Folha1"]
    res = {s: load_results(f"au_{s}", YEAR) for s in ("cm", "am", "af")}

    # slugs de todas as siglas conhecidas nos resultados (globais + por âmbito)
    global_slugs = set()
    for s in ("cm", "am", "af"):
        for k in res[s].get("METADATA", {}).get("parties", {}):
            global_slugs.add(strip_accents_upper(k))
        for agg in res[s].get("AGG", {}).get("concelho", {}).values():
            for k in (agg.get("votes") or {}):
                global_slugs.add(strip_accents_upper(k))
    for votes in res["af"].get("RESULTS", {}).values():
        for k in votes:
            global_slugs.add(strip_accents_upper(k))

    out = {"cm": {}, "am": {}}
    out_af = defaultdict(dict)   # dd -> {dicofre: orgao}
    warn = []

    for cod, conc, freg, org, rows in raw_blocks(ws):
        code = norm_dicofre(cod)
        if not code:
            warn.append(f"código inválido {cod!r} ({conc}/{org})")
            continue
        listas = parse_listas(rows, global_slugs)
        if not listas:
            warn.append(f"{cod} {conc} {org}: sem nomes")
            continue
        if org in ("CM", "AM"):
            dico = code[:4]
            agg = res[org.lower()].get("AGG", {}).get("concelho", {}).get(dico, {})
            local = set(agg.get("votes") or {}) | set(agg.get("mandatos_p") or {})
            listas = canonicalize_siglas(listas, local)
            entry = {"nome": conc, "listas": listas}
            if org == "CM":
                votes = (res["cm"].get("AGG", {}).get("concelho", {})
                         .get(dico, {}).get("votes"))
                p = compute_presidente(listas, votes)
                if p:
                    entry["presidente"] = p
                else:
                    warn.append(f"CM {dico} {conc}: presidente não determinado")
            out[org.lower()][dico] = entry
        else:
            votes = res["af"].get("RESULTS", {}).get(code)
            listas = canonicalize_siglas(listas, set(votes or {}))
            entry = {"nome": freg or conc, "listas": listas}
            p = compute_presidente(listas, votes)
            if p:
                entry["presidente"] = p
            dd = circulo_from_dicofre(code) or code[:2]
            out_af[dd][code] = entry

    for sub in ("cm", "am"):
        write_eleitos_json(f"au_{sub}_{YEAR}.json", {
            "year": YEAR, "election": "au", "subtype": sub, "orgaos": out[sub]})
        print(f"au_{sub}_{YEAR}: {len(out[sub])} concelhos")
    total_f = 0
    for dd, orgaos in sorted(out_af.items()):
        write_eleitos_json(f"au_af_{YEAR}_{dd}.json", {
            "year": YEAR, "election": "au", "subtype": "af", "distrito": dd,
            "orgaos": orgaos})
        total_f += len(orgaos)
    print(f"au_af_{YEAR}: {total_f} freguesias em {len(out_af)} distritos")
    for w in warn[:40]:
        print(f"  aviso: {w}")
    if len(warn) > 40:
        print(f"  ... +{len(warn) - 40} avisos")
    rebuild_index()


if __name__ == "__main__":
    sys.exit(main())
